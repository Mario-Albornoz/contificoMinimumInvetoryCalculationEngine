import openpyxl

from databaseConnector import databaseManager
from product import ProductData

db = databaseManager()
db.initialize_schema()
file_path = 'files/ReporteSaldosDisponibles.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active


def gather_data():
    date_string = get_value_from_sheet('Rango de Fechas', ws)
    bodega_string = get_value_from_sheet('Bodega', ws)

    if date_string:
        start_date, end_date = parse_date_string(date_string)
    else:
        print("Date range not found in the spreadsheet")

    if bodega_string:
        bodega = bodega_string.split(':')[1].strip()
    else:
        print('Bodega not found in spreadsheet')

    headers = [cell.value for cell in ws[6]]
    products = []
    for row in ws.iter_rows(min_row = 7, values_only = True):
        product_code_idx = headers.index('Código')
        product_name_idx = headers.index('Nombre')
        inicial_stock_idx = headers.index('Inicial')
        final_stock_idx = headers.index('Stock Final')
        unit_type_idx = headers.index('Unidad')

        product = ProductData(
            product_code=row[product_code_idx],
            product_name = row[product_name_idx],
            inicial_stock = row[inicial_stock_idx],
            final_stock = row[final_stock_idx],
            unit_type=row[unit_type_idx]
        )
        if product.unit_type is None:
            print('-------------------------------------------------------------------------')
            print("UNIDAD NULL: ", product)
        print(product)

        products.append(product)

    period_id = db.insert_period_record(start_date, end_date, warehouse=bodega)

    for product in products:
        product_id = db.upsert_product(product.product_name, product.product_code, unit_type=product.unit_type)
        db.insert_inventory_record(product_id, period_id, product.inicial_stock, product.final_stock)

    print("dataset generated sucessfuly")


def get_value_from_sheet(value_string:str, worksheet):
    found_value = None
    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=10):
        for cell in row:
            if cell.value and value_string in str(cell.value):
                found_value = str(cell.value)
                break
        if found_value:
            break

    return found_value

def parse_date_string(date_string):
    date_part = date_string.split(':')[1].strip()
    dates = date_part.split(' - ')
    start_date = dates[0].strip()
    end_date = dates[1].strip()

    return start_date, end_date

gather_data()
db.close()