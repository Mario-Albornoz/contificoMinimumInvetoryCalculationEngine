from datetime import datetime

import openpyxl

from modules.product import ProductData

def extract_products_from_report(ws) -> list:
    headers = [cell.value for cell in ws[6]]
    products = []
    for row in ws.iter_rows(min_row=7, values_only=True):
        product_code_idx = headers.index('Código')
        product_name_idx = headers.index('Nombre')
        initial_stock_idx = headers.index('Inicial')
        final_stock_idx = headers.index('Stock Final')
        unit_type_idx = headers.index('Unidad')

        product = ProductData(
            product_code=row[product_code_idx],
            product_name=row[product_name_idx],
            initial_stock=row[initial_stock_idx],
            final_stock=row[final_stock_idx],
            unit_type=row[unit_type_idx]
        )

        products.append(product)
    return products

def get_value_from_sheet(value_string:str, ws):
    found_value = None
    for row in ws.iter_rows(min_row=1, max_row=5, min_col=1, max_col=10):
        for cell in row:
            if cell.value and value_string in str(cell.value):
                found_value = str(cell.value)
                break
        if found_value:
            break
    return found_value

def parse_date_string(date_string):
    date_part = date_string.split(':')[1].strip()
    dates = date_part.split(' - ')
    start_date = dates[0].strip()
    end_date = dates[1].strip()

    return start_date, end_date

def gather_data_from_report(warehouse_id:str, current_file_path:str, db):
    ws = set_current_workbook(current_file_path)

    date_string = get_value_from_sheet('Rango de Fechas', ws)
    products = extract_products_from_report(ws)

    if date_string:
        start_date, end_date = parse_date_string(date_string)
    else:
        print("Date range not found in the spreadsheet")

    period_id = db.insert_period_record(start_date, end_date, warehouse_id=warehouse_id)

    for product in products:
        product_id = db.upsert_product(product.product_name, product.product_code, unit_type=product.unit_type)
        db.insert_inventory_record(product_id, period_id, product.initial_stock, product.final_stock)

    print("Inserted report Successfully")

def parse_date(date: datetime) -> str:
    return date.strftime("%d%%2F%m%%2F%Y")

def set_current_workbook(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    return ws

